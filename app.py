import streamlit as st
import pandas as pd
import json
import os
import shutil
import base64
import io
import qrcode
import requests
from datetime import datetime
import calendar
import openpyxl
from streamlit.components.v1 import html as st_html
from mercado_pago import criar_pix, consultar_pagamento, ultimo_status, TOKEN_CONFIGURADO, buscar_por_referencia
import threading
import time

st.set_page_config(page_title="Gestão de Entregas 2.0", layout="wide")

# ========== CONFIGURACOES DE PAGAMENTO ==========
VALOR_PIX = 50.00
# Chave PIX: CPF 36785517850 — Diego

# Funcao para gerar BR Code PIX valido
def gerar_br_code_pix(chave: str, valor: float, nome: str, cidade: str) -> str:
    """Gera um BR Code PIX valido segundo o Bacen."""
    gui = "BR.GOV.BCB.PIX"
    mai = f"0014{gui}01{len(chave):02d}{chave}"
    
    payload = f"00020126{len(mai):02d}{mai}"
    payload += "52040000"  # Merchant Category Code
    payload += "5303986"   # Transaction Currency (BRL)
    
    valor_str = f"{valor:.2f}"
    payload += f"54{len(valor_str):02d}{valor_str}"
    
    payload += "5802BR"    # Country Code
    payload += f"59{len(nome):02d}{nome}"
    payload += f"60{len(cidade):02d}{cidade}"
    
    txid = "***"
    adf = f"05{len(txid):02d}{txid}"
    payload += f"62{len(adf):02d}{adf}"
    
    payload += "6304"
    
    def crc16_ccitt(data: str) -> str:
        crc = 0xFFFF
        for byte in data.encode():
            crc ^= byte << 8
            for _ in range(8):
                if crc & 0x8000:
                    crc = (crc << 1) ^ 0x1021
                else:
                    crc <<= 1
            crc &= 0xFFFF
        return f"{crc:04X}"
    
    crc = crc16_ccitt(payload)
    return payload + crc

# Gera o BR Code PIX valido
PIX_BR_CODE = gerar_br_code_pix("36785517850", VALOR_PIX, "Diego", "SAOPAULO")

def criar_nova_planilha(caminho):
    """Cria uma planilha zerada com todas as datas dos meses."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    meses_info = [
        ("Junho", 6), ("Julho", 7), ("Agosto", 8), ("Setembro", 9),
        ("Outubro", 10), ("Novembro", 11), ("Dezembro", 12),
    ]
    for nome_mes, mes_num in meses_info:
        ws = wb.create_sheet(title=nome_mes)
        _, ultimo_dia = calendar.monthrange(2026, mes_num)
        for dia in range(1, ultimo_dia + 1):
            dt = datetime(2026, mes_num, dia)
            r = 3 + dia - 1
            ws.cell(row=r, column=2, value=dt)
    wb.save(caminho)
    wb.close()


def garantir_planilha_usuario(caminho):
    if os.path.exists(caminho):
        return True
    criar_nova_planilha(caminho)
    return True


# ========== CONFIGURACOES DE PAGAMENTO ==========
VALOR_PIX = 50.00
PIX_CHAVE = "00020126580014BR.GOV.BCB.PIX0136diego@seudominio.com5204000053039865404{valor}5802BR5925Diego Gestao Entregas 2.06009SAOPAULO62070503***6304"  # Pix Copia-e-Cola de exemplo — SUBSTITUA PELO SEU

def gerar_qrcode_pix(valor: float, chave: str = PIX_CHAVE, descricao: str = "Acesso Planilha de Ganhos"):
    """Gera QR Code do PIX e retorna base64 da imagem."""
    import qrcode as _qr
    buf = io.BytesIO()
    img = _qr.make(chave)
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{b64}"


import random
import string


def gerar_codigo_liberacao(tamanho=8):
    """Gera um codigo aleatorio de letras e numeros."""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=tamanho))


def salvar_codigo_pendente(login: str, codigo: str):
    """Salva o codigo de liberacao no arquivo de usuarios."""
    usuarios = carregar_usuarios()
    login = login.strip().lower()
    if login in usuarios:
        usuarios[login]["codigo_liberacao"] = codigo
        salvar_usuarios(usuarios)


def validar_codigo_liberacao(login: str, codigo: str):
    """Verifica se o codigo informado esta correto."""
    usuarios = carregar_usuarios()
    login = login.strip().lower()
    if login in usuarios:
        return usuarios[login].get("codigo_liberacao", "") == codigo.strip().upper()
    return False


def marcar_usuario_pago(login: str):
    usuarios = carregar_usuarios()
    login = login.strip().lower()
    if login in usuarios:
        usuarios[login]["pago"] = True
        usuarios[login].pop("codigo_liberacao", None)
        salvar_usuarios(usuarios)


# ========== LOGIN SISTEMA ==========
USUARIOS_JSON = "usuarios.json"
USUARIOS_APPEND = "usuarios_append.jsonl"  # Arquivo append-only (uma linha por usuario)

def _ler_append():
    """Le o arquivo append-only e retorna dict de usuarios."""
    usuarios = {}
    if os.path.exists(USUARIOS_APPEND):
        try:
            with open(USUARIOS_APPEND, "r", encoding="utf-8") as f:
                for linha in f:
                    linha = linha.strip()
                    if not linha:
                        continue
                    try:
                        registro = json.loads(linha)
                        login = registro.get("login", "").strip().lower()
                        if login:
                            usuarios[login] = registro.get("dados", {})
                    except Exception:
                        continue
        except Exception:
            pass
    return usuarios

def _append_usuario(login: str, dados: dict):
    """Adiciona um usuario ao arquivo append-only (nunca sobrescreve)."""
    registro = {
        "login": login,
        "dados": dados,
        "timestamp": datetime.now().isoformat()
    }
    with open(USUARIOS_APPEND, "a", encoding="utf-8") as f:
        f.write(json.dumps(registro, ensure_ascii=False) + "\n")
        f.flush()
        os.fsync(f.fileno())

def _reconstruir_json_principal():
    """Reconstroi o usuarios.json a partir do append + json existente."""
    # Le o JSON principal (se existir)
    principal = {}
    if os.path.exists(USUARIOS_JSON):
        try:
            with open(USUARIOS_JSON, "r", encoding="utf-8") as f:
                principal = json.load(f)
        except Exception:
            principal = {}
    
    # Le o append (novos usuarios)
    append = _ler_append()
    
    # Merge: append sobrescreve principal (usuarios mais recentes ganham)
    merge = {**principal, **append}
    
    # Salva o merge de volta no JSON principal
    with open(USUARIOS_JSON, "w", encoding="utf-8") as f:
        json.dump(merge, f, indent=4, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    
    return merge

def carregar_usuarios():
    """Carrega usuarios do JSON, reconstruindo a partir do append se necessario."""
    # Sempre reconstrui a partir do append para garantir consistencia
    return _reconstruir_json_principal()

def salvar_usuarios(usuarios: dict):
    """Salva todos os usuarios no JSON principal (usado por admin)."""
    with open(USUARIOS_JSON, "w", encoding="utf-8") as f:
        json.dump(usuarios, f, indent=4, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())


def criar_usuario(login: str, nome: str, senha: str, telefone: str = ""):
    login = login.strip().lower()
    nome = nome.strip()
    senha = senha.strip()
    telefone = telefone.strip()
    if not login or not nome or not senha:
        st.error("Preencha todos os campos obrigatórios.")
        return
    if len(senha) < 4:
        st.error("Senha muito curta.")
        return
    usuarios = carregar_usuarios()
    if login in usuarios:
        st.error("Usuário já existe.")
        return
    planilha_nome = f"Planilha de Ganhos - {nome.title()}.xlsx"
    dados_usuario = {
        "nome": nome,
        "senha": senha,
        "telefone": telefone,
        "planilha": planilha_nome,
        "pago": False
    }
    # Salva no append-only (nunca perde dados)
    _append_usuario(login, dados_usuario)
    # Atualiza o JSON principal
    usuarios[login] = dados_usuario
    salvar_usuarios(usuarios)
    garantir_planilha_usuario(planilha_nome)
    st.success(f"Conta criada! Usuario {login} salvo com sucesso.")
    # Redireciona para tela de pagamento
    st.session_state["tela"] = "pagamento"
    st.session_state["usuario_pendente"] = {**dados_usuario, "login": login}
    st.rerun()


def logout():
    for chave in ["logado", "usuario", "arquivo_excel", "dados", "admin_logado"]:
        st.session_state.pop(chave, None)
    st.rerun()


def tela_admin():
    """Tela de gerenciamento interno de usuarios (apenas admin)."""
    st.markdown("""
    <style>
    .admin-header { color: #fff; font-size: 28px; font-weight: 800; margin-bottom: 24px; text-align: center; }
    .admin-card { background: rgba(255,255,255,0.05); border-radius: 16px; padding: 24px; border: 1px solid rgba(255,255,255,0.1); }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:4vh'></div>", unsafe_allow_html=True)
    st.markdown("<div class='admin-header'>Painel Administrativo</div>", unsafe_allow_html=True)

    usuarios = carregar_usuarios()

    # Estatisticas
    total = len(usuarios)
    pagos = sum(1 for u in usuarios.values() if u.get("pago", False))
    pendentes = total - pagos

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total de Usuários", total)
    with col2:
        st.metric("Pagos", pagos)
    with col3:
        st.metric("Pendentes", pendentes)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # === USUARIOS PENDENTES ===
    if pendentes > 0:
        st.subheader("&#x1f514; Usuários Pendentes de Pagamento")
        pendentes_list = [(login, info) for login, info in usuarios.items() if not info.get("pago", False)]
        for login, info in pendentes_list:
            with st.container():
                c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
                with c1:
                    st.markdown(f"**<span style='color:#fbbf24'>{login}</span>**")
                with c2:
                    st.caption(info.get("nome", ""))
                with c3:
                    st.caption(info.get("planilha", ""))
                with c4:
                    if st.button("Liberar", key=f"liberar_{login}", type="primary"):
                        marcar_usuario_pago(login)
                        st.success(f"Usuário **{login}** liberado!")
                        st.rerun()
        st.markdown("---")

    # === TABELA DE TODOS OS USUARIOS ===
    st.subheader("Lista de Usuários Cadastrados")

    dados_tabela = []
    for login, info in usuarios.items():
        dados_tabela.append({
            "Login": login,
            "Nome": info.get("nome", ""),
            "Senha": info.get("senha", ""),
            "Pago": "Sim" if info.get("pago", False) else "Não",
            "Planilha": info.get("planilha", ""),
        })

    df_usuarios = pd.DataFrame(dados_tabela)
    st.dataframe(df_usuarios, use_container_width=True, hide_index=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # Acoes por usuario
    st.subheader("Gerenciar Usuário")
    col_sel, col_acoes = st.columns([1, 2])

    with col_sel:
        usuario_sel = st.selectbox("Selecione o usuário", list(usuarios.keys()))

    if usuario_sel:
        user_data = usuarios[usuario_sel]
        with col_acoes:
            c1, c2, c3 = st.columns(3)

            with c1:
                if not user_data.get("pago", False):
                    if st.button("Marcar como Pago", use_container_width=True, type="primary"):
                        marcar_usuario_pago(usuario_sel)
                        st.success(f"Usuário **{usuario_sel}** liberado!")
                        st.rerun()
                else:
                    if st.button("Bloquear Acesso", use_container_width=True):
                        usuarios[usuario_sel]["pago"] = False
                        salvar_usuarios(usuarios)
                        st.warning(f"Usuário **{usuario_sel}** bloqueado.")
                        st.rerun()

            with c2:
                if st.button("Excluir Usuário", use_container_width=True):
                    del usuarios[usuario_sel]
                    salvar_usuarios(usuarios)
                    st.error(f"Usuário **{usuario_sel}** excluído.")
                    st.rerun()

            with c3:
                if st.button("Ver Dados", use_container_width=True):
                    st.json(user_data)

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    # Adicionar novo usuario manualmente
    with st.expander("Adicionar Novo Usuário"):
        col_nome, col_user, col_senha = st.columns(3)
        with col_nome:
            novo_nome = st.text_input("Nome Completo", key="admin_novo_nome")
        with col_user:
            novo_login = st.text_input("Login", key="admin_novo_login")
        with col_senha:
            nova_senha = st.text_input("Senha", type="password", key="admin_nova_senha")

        col_pago, col_espaco = st.columns([1, 3])
        with col_pago:
            novo_pago = st.checkbox("Já está pago?", key="admin_novo_pago")

        if st.button("Criar Usuário", type="primary"):
            if novo_login and novo_nome and nova_senha:
                criar_usuario(novo_login, novo_nome, nova_senha)
                if novo_pago:
                    marcar_usuario_pago(novo_login)
                st.success("Usuário criado com sucesso!")
                st.rerun()
            else:
                st.error("Preencha todos os campos.")

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    if st.button("Sair do Painel", use_container_width=True):
        st.session_state.pop("admin_logado", None)
        st.rerun()


def tela_pagamento():
    user = st.session_state.get("usuario_pendente", {})
    login_pendente = user.get("login", "")
    nome_pendente = user.get("nome", "")

    # Se ja foi pago (outra aba/admin)
    usuarios = carregar_usuarios()
    if usuarios.get(login_pendente, {}).get("pago", False):
        st.session_state["logado"] = True
        st.session_state["usuario"] = usuarios[login_pendente]
        st.session_state["arquivo_excel"] = usuarios[login_pendente]["planilha"]
        st.session_state.pop("usuario_pendente", None)
        st.rerun()

    # Gera QR Code estatico da chave pessoal (CPF) — sempre o mesmo
    if "qr_pix_b64" not in st.session_state:
        buf = io.BytesIO()
        img = qrcode.make(PIX_BR_CODE)
        img.save(buf, format="PNG")
        st.session_state["qr_pix_b64"] = base64.b64encode(buf.getvalue()).decode("utf-8")
    qr_b64 = st.session_state["qr_pix_b64"]

    _, c2, _ = st.columns([1, 2.6, 1])
    with c2:
        st.markdown("<div style='height:6vh'></div>", unsafe_allow_html=True)
        st.markdown("<div style='text-align:center;'><span style='font-size:48px;'>&#x1f512;</span></div>", unsafe_allow_html=True)
        st.markdown("<h2 style='color:#fff;text-align:center;font-weight:800;letter-spacing:2px;text-transform:uppercase;margin-top:4px;'>Acesso Bloqueado</h2>", unsafe_allow_html=True)
        st.markdown("<p style='color:#a5b4fc;text-align:center;font-size:15px;margin-bottom:24px;'>Efetue o pagamento PIX de <b>R&#36; 20,00</b> para liberar seu acesso.</p>", unsafe_allow_html=True)

        # QR Code
        st.markdown(f'<div style="text-align:center;margin-bottom:16px;"><img src="data:image/png;base64,{qr_b64}" style="width:240px;border-radius:12px;box-shadow:0 8px 24px rgba(0,0,0,0.4);"></div>', unsafe_allow_html=True)

        # Copia e Cola
        st.markdown("<p style='color:#9ca3af;text-align:center;font-size:12px;margin-bottom:8px;'>Cole no app do seu banco</p>", unsafe_allow_html=True)
        st.markdown(f'<div style="background:rgba(0,0,0,0.25);border-radius:12px;padding:12px;border:1px dashed rgba(255,255,255,0.15);text-align:center;margin-bottom:14px;overflow-wrap:break-word;"><span style="color:#111;background:#fff;padding:6px 10px;border-radius:4px;font-size:13px;font-weight:700;">{PIX_BR_CODE}</span></div>', unsafe_allow_html=True)

        st.markdown("<p style='color:#fbbf24;text-align:center;font-size:12px;margin:10px 0;'>Após o pagamento, envie o comprovante pelo WhatsApp e aguarde a liberação pelo administrador.</p>", unsafe_allow_html=True)

        if st.button("Voltar ao Login", use_container_width=True):
            st.session_state["tela"] = "login"
            st.session_state.pop("usuario_pendente", None)
            st.session_state.pop("qr_pix_b64", None)
            st.rerun()


def tela_login_admin():
    """Tela de login exclusiva para o painel administrativo."""
    st.markdown("""
    <style>
    .stApp { background: #0f0f1a !important; }
    .main { background: #0f0f1a !important; }
    </style>
    """, unsafe_allow_html=True)

    _, c2, _ = st.columns([1, 2, 1])
    with c2:
        st.markdown("<div style='height:15vh'></div>", unsafe_allow_html=True)
        st.markdown("<div style='text-align:center;'><span style='font-size:48px;'>&#x1f510;</span></div>", unsafe_allow_html=True)
        st.markdown("<h2 style='color:#fff;text-align:center;font-weight:800;letter-spacing:2px;text-transform:uppercase;margin-top:4px;'>Painel Administrativo</h2>", unsafe_allow_html=True)
        st.markdown("<p style='color:#a5b4fc;text-align:center;font-size:15px;margin-bottom:24px;'>Acesso restrito para administradores.</p>", unsafe_allow_html=True)

        senha_admin = st.text_input("Senha de Administrador", type="password", placeholder="Digite a senha", key="admin_senha_input", label_visibility="collapsed")

        if st.button("Entrar no Painel", type="primary", use_container_width=True):
            if senha_admin == "admin2026":
                st.session_state["admin_logado"] = True
                st.rerun()
            else:
                st.error("Senha incorreta.")

        if st.button("Voltar ao Login", use_container_width=True):
            st.session_state["tela"] = "login"
            st.rerun()


def tela_login():
    if "tela" not in st.session_state:
        st.session_state["tela"] = "login"

    # CSS global + orbes
    st.markdown("""
    <style>
    .stApp { background: #0f0f1a !important; }
    .main { background: #0f0f1a !important; }
    .block-container { padding-top: 0 !important; padding-bottom: 0 !important; }
    .orb {
        position: fixed; border-radius: 50%; filter: blur(120px);
        z-index: -1; pointer-events: none;
    }
    .orb-1 { width: 340px; height: 340px; background: #7c3aed; top: -80px; left: -80px; opacity: 0.5; }
    .orb-2 { width: 300px; height: 300px; background: #db2777; bottom: -80px; right: -80px; opacity: 0.5; }
    .orb-3 { width: 220px; height: 220px; background: #f59e0b; top: 50%%; left: 55%%; opacity: 0.3; }
    input[type="text"], input[type="password"] {
        background: rgba(255,255,255,0.85) !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 12px !important; color: #000000 !important;
        padding: 12px 14px !important; font-size: 14px !important;
    }
    input[type="text"]::placeholder, input[type="password"]::placeholder {
        color: #888888 !important;
    }
    button[kind="primary"] {
        background: rgba(255,255,255,0.92) !important; color: #111 !important;
        border: none !important; border-radius: 50px !important;
        font-weight: 700 !important; font-size: 14px !important;
        letter-spacing: 1.5px; text-transform: uppercase;
        padding: 12px 28px !important;
        box-shadow: 0 8px 24px rgba(255,255,255,0.12) !important;
    }
    button[kind="primary"]:hover {
        background: #fff !important;
    }
    button[kind="secondary"] {
        background: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(255,255,255,0.20) !important;
        border-radius: 50px !important;
        color: #fff !important;
        font-weight: 600 !important; font-size: 13px !important;
        letter-spacing: 1px; text-transform: uppercase;
        padding: 10px 28px !important;
    }
    button[kind="secondary"]:hover {
        background: rgba(255,255,255,0.16) !important;
        border-color: rgba(255,255,255,0.35) !important;
    }
    </style>
    <div class="orb orb-1"></div>
    <div class="orb orb-2"></div>
    <div class="orb orb-3"></div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:10vh'></div>", unsafe_allow_html=True)

    # Card glassmorphism puro HTML/Div
    st.markdown("""
    <div style="max-width:780px;margin:0 auto;background:rgba(255,255,255,0.04);backdrop-filter:blur(20px);border-radius:24px;border:1px solid rgba(255,255,255,0.10);box-shadow:0 40px 80px rgba(0,0,0,0.5);">
        <div style="display:flex;flex-wrap:wrap;">
            <div style="flex:1;min-width:260px;background:rgba(255,255,255,0.03);border-right:1px solid rgba(255,255,255,0.07);padding:48px 36px;display:flex;flex-direction:column;justify-content:center;">
                <div style="color:#fff;font-size:22px;font-weight:800;letter-spacing:3px;margin-bottom:12px;text-transform:uppercase;">Bem-vindo</div>
                <div style="color:#a5b4fc;font-size:13px;margin-bottom:40px;line-height:1.5;">Acesse sua planilha de ganhos diários ou crie uma nova conta para começar.</div>
                <div id="btn-area-left"></div>
            </div>
            <div style="flex:1.25;min-width:300px;padding:48px 36px;display:flex;flex-direction:column;justify-content:center;">
                <div id="form-area-right"></div>
            </div>
        </div>
    </div>
    <div style='height:8vh'></div>
    """, unsafe_allow_html=True)

    # Conteudo real interativo
    c1, c2, c3 = st.columns([1, 5.5, 1])
    with c2:
        if st.session_state.get("tela") == "pagamento":
            tela_pagamento()
            return

        col_left, col_right = st.columns([1, 1.25])

        with col_left:
            st.markdown("<div style='height:22px'></div>", unsafe_allow_html=True)
            if st.session_state["tela"] == "login":
                st.markdown("<div style='color:#9ca3af;font-size:11px;margin-bottom:10px;'>Ainda não tem conta?</div>", unsafe_allow_html=True)
                if st.button("Criar Conta", key="btn_criar"):
                    st.session_state["tela"] = "cadastro"
                    st.rerun()
                st.markdown("<div style='color:#9ca3af;font-size:11px;margin-top:14px;margin-bottom:10px;'>Acesso administrativo</div>", unsafe_allow_html=True)
                if st.button("Painel Admin", key="btn_admin"):
                    st.session_state["tela"] = "admin"
                    st.rerun()
            else:
                st.markdown("<div style='color:#9ca3af;font-size:11px;margin-bottom:10px;'>Já tem conta?</div>", unsafe_allow_html=True)
                if st.button("Fazer Login", key="btn_login"):
                    st.session_state["tela"] = "login"
                    st.rerun()

        with col_right:
            if st.session_state["tela"] == "login":
                st.markdown("<div style='color:#fff;font-size:20px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:24px;text-align:center;'>Faça Login</div>", unsafe_allow_html=True)
                login = st.text_input("Usuário", placeholder="usuario", key="login_user", label_visibility="collapsed")
                senha = st.text_input("Senha", type="password", placeholder="senha", key="login_senha", label_visibility="collapsed")
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                usuarios = carregar_usuarios()
                if st.button("Entrar", type="primary", use_container_width=True, key="btn_entrar"):
                    user = usuarios.get(login.strip().lower())
                    if user and user["senha"] == senha:
                        if not user.get("pago", False):
                            st.session_state["tela"] = "pagamento"
                            st.session_state["usuario_pendente"] = {**user, "login": login.strip().lower()}
                            st.rerun()
                        st.session_state["logado"] = True
                        st.session_state["usuario"] = user
                        st.session_state["arquivo_excel"] = user["planilha"]
                        st.rerun()
                    else:
                        st.error("Usuário ou senha incorretos.")
                st.markdown("<div style='color:#9ca3af;font-size:10px;text-align:center;margin-top:12px;'>Gestão de Entregas 2.0 &bull; 2026</div>", unsafe_allow_html=True)
            else:
                st.markdown("<div style='color:#fff;font-size:20px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:24px;text-align:center;'>Criar Conta</div>", unsafe_allow_html=True)
                novo_nome = st.text_input("Nome Completo", placeholder="nome completo", key="cad_nome", label_visibility="collapsed")
                novo_telefone = st.text_input("Telefone", placeholder="55xx xxxxx-xxxx", key="cad_telefone", label_visibility="collapsed")
                novo_user = st.text_input("Usuário", placeholder="usuario", key="cad_user", label_visibility="collapsed")
                nova_senha = st.text_input("Senha", type="password", placeholder="senha", key="cad_senha", label_visibility="collapsed")
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                if st.button("Cadastrar", type="primary", use_container_width=True, key="btn_cadastrar"):
                    criar_usuario(novo_user, novo_nome, nova_senha, novo_telefone)


if "admin_logado" in st.session_state and st.session_state["admin_logado"]:
    tela_admin()
    st.stop()

if "logado" not in st.session_state or not st.session_state["logado"]:
    # Verifica se é login admin
    if st.session_state.get("tela") == "admin":
        tela_login_admin()
        st.stop()
    tela_login()
    st.stop()

ARQUIVO_EXCEL = st.session_state["arquivo_excel"]

# ========== FIM LOGIN ==========

# Tabela de lookup da diaria (igual ao VLOOKUP do Excel)
DIARIA_LOOKUP = [
    (0, 240), (101, 260), (126, 280), (151, 300), (176, 320),
    (201, 340), (226, 360), (251, 380), (276, 400), (301, 420),
    (326, 440), (351, 460), (376, 480)
]


def calc_diaria(km, is_sunday: bool):
    if pd.isna(km):
        return None
    try:
        km = float(km)
    except (ValueError, TypeError):
        return None
    if km <= 0:
        return 0
    val = 0
    for thr, v in DIARIA_LOOKUP:
        if km >= thr:
            val = v
        else:
            break
    if is_sunday:
        val += 48
    return val


def calc_pct(pacotes, entregas):
    try:
        p = float(pacotes)
        e = float(entregas)
        if p == 0:
            return None
        return e / p
    except (TypeError, ValueError):
        return None


def calc_bonus(pct):
    if pct is None:
        return 0
    if pct == 1.0:
        return 20
    if 0.98 <= pct < 1.0:
        return 10
    return 0


def dia_semana_br(dt):
    if pd.isna(dt):
        return ""
    dias = [
        "segunda-feira", "terca-feira", "quarta-feira", "quinta-feira",
        "sexta-feira", "sabado", "domingo"
    ]
    return dias[dt.weekday()]


MESES = {
    "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}


def importar_excel_original():
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
    dados = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        mes = sheet.strip().lower()
        if mes not in MESES:
            continue
        rows = []
        empties = 0
        for r in range(4, 100):
            v_data = ws.cell(row=r, column=2).value
            if v_data is None:
                empties += 1
                if empties >= 10:
                    break
                continue
            empties = 0
            if isinstance(v_data, str):
                try:
                    v_data = datetime.strptime(v_data, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    continue
            rows.append({
                "data": v_data,
                "km": ws.cell(row=r, column=4).value,
                "pacotes": ws.cell(row=r, column=6).value,
                "entregas": ws.cell(row=r, column=7).value,
                "gasolina": ws.cell(row=r, column=15).value,
                "almoco": ws.cell(row=r, column=16).value,
                "besteira": ws.cell(row=r, column=17).value,
                "carro": ws.cell(row=r, column=18).value,
            })
        df = pd.DataFrame(rows)
        for c in ["km", "pacotes", "entregas", "gasolina", "almoco", "besteira", "carro"]:
            if c not in df.columns:
                df[c] = None
        if df.empty:
            df = pd.DataFrame(columns=["data", "km", "pacotes", "entregas",
                                        "gasolina", "almoco", "besteira", "carro"])
        dados[mes] = df
    wb.close()
    return dados


def criar_vazio_para_mes(ano: int, mes_num: int):
    _, last_day = calendar.monthrange(ano, mes_num)
    dias = [datetime(ano, mes_num, d) for d in range(1, last_day + 1)]
    df = pd.DataFrame({
        "data": dias,
        "km": [None] * len(dias),
        "pacotes": [None] * len(dias),
        "entregas": [None] * len(dias),
        "gasolina": [None] * len(dias),
        "almoco": [None] * len(dias),
        "besteira": [None] * len(dias),
        "carro": [None] * len(dias),
    })
    return df


def ler_planilha():
    try:
        return importar_excel_original()
    except Exception as e:
        st.toast(f"Erro ao importar Excel: {e}", icon="⚠️")
        return {}


def inicializar_dados():
    if "dados" not in st.session_state:
        dados_importados = ler_planilha()
        dados = {}
        for mes, num in MESES.items():
            if mes in dados_importados and not dados_importados[mes].empty:
                dados[mes] = dados_importados[mes].copy()
            else:
                dados[mes] = criar_vazio_para_mes(2026, num)
        st.session_state["dados"] = dados


def calcular_df(df):
    out = df.copy()
    for c in ["km", "pacotes", "entregas", "gasolina", "almoco", "besteira", "carro"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    out["dia_semana"] = out["data"].apply(dia_semana_br)
    out["is_sunday"] = out["data"].apply(lambda d: d.weekday() == 6 if pd.notna(d) else False)
    out["diaria"] = out.apply(lambda r: calc_diaria(r["km"], r["is_sunday"]), axis=1)
    out["pct_entrega"] = out.apply(lambda r: calc_pct(r["pacotes"], r["entregas"]), axis=1)
    out["bonus"] = out["pct_entrega"].apply(calc_bonus)
    diaria_num = pd.to_numeric(out["diaria"], errors="coerce").fillna(0)
    bonus_num = pd.to_numeric(out["bonus"], errors="coerce").fillna(0)
    out["total_ganho"] = diaria_num + bonus_num
    out["total_gasto"] = (
        pd.to_numeric(out["gasolina"], errors="coerce").fillna(0) +
        pd.to_numeric(out["almoco"], errors="coerce").fillna(0) +
        pd.to_numeric(out["besteira"], errors="coerce").fillna(0) +
        pd.to_numeric(out["carro"], errors="coerce").fillna(0)
    )
    out["lucro"] = out["total_ganho"] - out["total_gasto"]
    return out


def resumo(df_calc):
    # 1a quinzena: dia 26 ao 10
    # 2a quinzena: dia 11 ao 25
    df_calc = df_calc.copy()
    df_calc["dia"] = df_calc["data"].apply(lambda d: d.day if pd.notna(d) else None)

    def _is_q1(dia):
        if dia is None:
            return False
        return dia >= 26 or dia <= 10

    def _is_q2(dia):
        if dia is None:
            return False
        return 11 <= dia <= 25

    df1 = df_calc[df_calc["dia"].apply(_is_q1)]
    df2 = df_calc[df_calc["dia"].apply(_is_q2)]

    pac1 = df1["pacotes"].sum()
    ent1 = df1["entregas"].sum()
    tax1 = ent1 / pac1 if pac1 else None
    dias1 = int(df1["km"].notna().sum())
    g1 = df1["total_ganho"].sum()
    b600_1 = 600 if (dias1 >= 13 and tax1 is not None and 0.98 <= tax1 <= 1.0) else 0
    t1 = g1 + b600_1

    pac2 = df2["pacotes"].sum()
    ent2 = df2["entregas"].sum()
    tax2 = ent2 / pac2 if pac2 else None
    dias2 = int(df2["km"].notna().sum())
    g2 = df2["total_ganho"].sum()
    b600_2 = 600 if (dias2 >= 13 and tax2 is not None and 0.98 <= tax2 <= 1.0) else 0
    t2 = g2 + b600_2

    total_g = t1 + t2
    total_s = df_calc["total_gasto"].sum()
    return {
        "ganho1": g1, "taxa1": tax1, "dias1": dias1, "bonus600_1": b600_1, "total1": t1,
        "ganho2": g2, "taxa2": tax2, "dias2": dias2, "bonus600_2": b600_2, "total2": t2,
        "total_ganho": total_g, "total_gasto": total_s, "total_lucro": total_g - total_s,
    }


def salvar_no_excel_original(dados_dict):
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=False)
    for sheet in wb.sheetnames:
        mes = sheet.strip().lower()
        if mes not in dados_dict:
            continue
        df = dados_dict[mes]
        ws = wb[sheet]
        data_para_linha = {}
        for r_excel in range(4, ws.max_row + 1):
            v_data = ws.cell(row=r_excel, column=2).value
            if isinstance(v_data, datetime):
                chave = v_data.strftime("%Y-%m-%d")
                data_para_linha[chave] = r_excel

        for _, row in df.iterrows():
            dt = row["data"]
            if pd.isna(dt):
                continue
            chave = dt.strftime("%Y-%m-%d")
            if chave not in data_para_linha:
                continue
            r = data_para_linha[chave]
            km = row["km"]
            ws.cell(row=r, column=4).value = int(km) if pd.notna(km) else None
            pac = row["pacotes"]
            ws.cell(row=r, column=6).value = int(pac) if pd.notna(pac) else None
            ent = row["entregas"]
            ws.cell(row=r, column=7).value = int(ent) if pd.notna(ent) else None
            gas = row["gasolina"]
            ws.cell(row=r, column=15).value = float(gas) if pd.notna(gas) else None
            alm = row["almoco"]
            ws.cell(row=r, column=16).value = float(alm) if pd.notna(alm) else None
            bes = row["besteira"]
            ws.cell(row=r, column=17).value = float(bes) if pd.notna(bes) else None
            car = row["carro"]
            ws.cell(row=r, column=18).value = float(car) if pd.notna(car) else None
    wb.save(ARQUIVO_EXCEL)
    wb.close()


# =================== MAIN ===================

# Garante que a planilha do usuario exista
garantir_planilha_usuario(ARQUIVO_EXCEL)

inicializar_dados()

# Header com logout
h1, h2 = st.columns([6, 1])
with h1:
    st.title("Gestão de Entregas 2.0")
with h2:
    usuario_nome = st.session_state["usuario"].get("nome", "Usuário")
    st.markdown(f"<div style='text-align:right; padding-top:18px;'><b>{usuario_nome}</b></div>", unsafe_allow_html=True)
    if st.button("Sair 🔒", use_container_width=True):
        logout()

# Mes atual como padrao
mes_atual = datetime.now().month
mes_nomes = list(MESES.keys())
mes_index_padrao = 0
for i, (nome, num) in enumerate(MESES.items()):
    if num == mes_atual:
        mes_index_padrao = i
        break

mes_sel = st.selectbox("Mes", mes_nomes, index=mes_index_padrao)

tab_lanc, tab_res = st.tabs(["Lancamentos", "Resultado Calculado"])

with tab_lanc:
    st.subheader(f"Lancamentos - {mes_sel.capitalize()}/2026")

    # Monta lista de datas do mes (apenas dias do mes selecionado)
    df_mes = st.session_state["dados"][mes_sel].copy()
    mes_num = MESES[mes_sel]
    hoje = datetime.now()

    datas_opcoes = []
    idx_padrao = 0
    for i, row in df_mes.iterrows():
        dt = row["data"]
        if pd.notna(dt) and dt.month == mes_num:
            dia_sem = dia_semana_br(dt)
            label = f"{dt.strftime('%d/%m/%Y')} - {dia_sem}"
            datas_opcoes.append((i, label))
            if dt.day == hoje.day and mes_num == hoje.month:
                idx_padrao = len(datas_opcoes) - 1

    if not datas_opcoes:
        st.warning("Nenhuma data encontrada para este mes.")
        st.stop()

    sel_combo = st.selectbox(
        "Selecione a data",
        range(len(datas_opcoes)),
        index=idx_padrao,
        format_func=lambda i: datas_opcoes[i][1],
        key=f"sel_{mes_sel}"
    )

    sel_idx = datas_opcoes[sel_combo][0]

    # Pega valores atuais dessa linha
    linha = df_mes.iloc[sel_idx]

    with st.form(key=f"form_{mes_sel}_{sel_idx}"):
        c1, c2 = st.columns(2)
        with c1:
            km = st.number_input("Km Executado", min_value=0, value=int(linha["km"]) if pd.notna(linha["km"]) else 0, step=1)
            pacotes = st.number_input("Qtd. Pacotes", min_value=0, value=int(linha["pacotes"]) if pd.notna(linha["pacotes"]) else 0, step=1)
            entregas = st.number_input("Qtd. Entregas", min_value=0, value=int(linha["entregas"]) if pd.notna(linha["entregas"]) else 0, step=1)
        with c2:
            gasolina = st.number_input("Gasolina", min_value=0.0, value=float(linha["gasolina"]) if pd.notna(linha["gasolina"]) else 0.0, step=0.01)
            almoco = st.number_input("Almoço", min_value=0.0, value=float(linha["almoco"]) if pd.notna(linha["almoco"]) else 0.0, step=0.01)
            besteira = st.number_input("Besteira", min_value=0.0, value=float(linha["besteira"]) if pd.notna(linha["besteira"]) else 0.0, step=0.01)
            carro = st.number_input("Carro", min_value=0.0, value=float(linha["carro"]) if pd.notna(linha["carro"]) else 0.0, step=0.01)

        submitted = st.form_submit_button("Salvar")
        if submitted:
            st.session_state["dados"][mes_sel].at[sel_idx, "km"] = km if km > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "pacotes"] = pacotes if pacotes > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "entregas"] = entregas if entregas > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "gasolina"] = gasolina if gasolina > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "almoco"] = almoco if almoco > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "besteira"] = besteira if besteira > 0 else None
            st.session_state["dados"][mes_sel].at[sel_idx, "carro"] = carro if carro > 0 else None
            salvar_no_excel_original(st.session_state["dados"])
            st.success("Dados salvos na planilha!")
            st.rerun()

with tab_res:
    st.subheader("Resultado Calculado")
    calc = calcular_df(st.session_state["dados"][mes_sel].copy())

    # Resumo
    r = resumo(calc)

    # Cards superiores customizados
    st.markdown("---")
    col1, col2, col3, col4, col5 = st.columns(5)

    card_style = """
    <div style="background-color: {bg}; padding: 20px; border-radius: 12px; text-align: center; box-shadow: 2px 2px 8px rgba(0,0,0,0.1); margin-bottom: 10px;">
        <div style="font-size: 28px; margin-bottom: 8px;">{icon}</div>
        <div style="color: #666; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px;">{label}</div>
        <div style="color: {fg}; font-size: 24px; font-weight: 700; margin-top: 4px;">{value}</div>
    </div>
    """

    with col1:
        st.markdown(card_style.format(
            bg="#FFF3E0", icon="📅", label="1ª Quinzena",
            fg="#E65100", value=f"R$ {r['total1']:,.0f}".replace(",", ".")
        ), unsafe_allow_html=True)

    with col2:
        st.markdown(card_style.format(
            bg="#E3F2FD", icon="📅", label="2ª Quinzena",
            fg="#1565C0", value=f"R$ {r['total2']:,.0f}".replace(",", ".")
        ), unsafe_allow_html=True)

    with col3:
        st.markdown(card_style.format(
            bg="#E8F5E9", icon="💰", label="Total de Ganhos",
            fg="#2E7D32", value=f"R$ {r['total_ganho']:,.0f}".replace(",", ".")
        ), unsafe_allow_html=True)

    with col4:
        st.markdown(card_style.format(
            bg="#FFEBEE", icon="💸", label="Total de Gastos",
            fg="#C62828", value=f"R$ {r['total_gasto']:,.2f}".replace(",", ".")
        ), unsafe_allow_html=True)

    with col5:
        lucro_bg = "#E8F5E9" if r['total_lucro'] >= 0 else "#FFEBEE"
        lucro_fg = "#2E7D32" if r['total_lucro'] >= 0 else "#C62828"
        st.markdown(card_style.format(
            bg=lucro_bg, icon="📊", label="Lucro Líquido",
            fg=lucro_fg, value=f"R$ {r['total_lucro']:,.2f}".replace(",", ".")
        ), unsafe_allow_html=True)

    # Tabela de detalhes
    st.markdown("---")

    # Formata para exibicao
    fmt = calc.copy()
    # Filtra apenas dias com dados
    fmt = fmt[fmt["km"].notna() & (fmt["km"] > 0)].reset_index(drop=True)

    # Substituir NaN/None em colunas de gasto para evitar "R$ nan"
    for col in ["gasolina", "almoco", "besteira", "carro"]:
        fmt[col] = fmt[col].fillna(0)

    def fmt_val(val, dec=0):
        if pd.isna(val):
            return ""
        return f"{val:.{dec}f}"

    def fmt_rs(val, dec=0):
        if pd.isna(val):
            return ""
        return f"R$ {val:,.{dec}f}".replace(",", "_").replace("_", ".").replace(".", ",", 1)

    # Monta tabela HTML customizada
    html = """
    <style>
    .tabela-ganhos { width: 100%; border-collapse: collapse; font-size: 13px; font-family: sans-serif; }
    .tabela-ganhos thead th {
        background-color: #1565C0;
        color: #ffffff;
        padding: 10px 6px;
        text-align: center;
        font-weight: 600;
        border: 1px solid #0d47a1;
    }
    .tabela-ganhos td {
        padding: 8px 6px;
        text-align: center;
        border-bottom: 1px solid #e0e0e0;
        border-right: 1px solid #e0e0e0;
    }
    .tabela-ganhos tr:nth-child(even) td { background-color: #f5f5f5; }
    .tabela-ganhos tr:hover td { background-color: #e3f2fd; }
    .tabela-ganhos .text-left { text-align: left; padding-left: 10px; }
    .tabela-ganhos .numero { font-weight: 600; white-space: nowrap; }
    .tabela-ganhos .ganho { background-color: #E8F5E9 !important; }
    .tabela-ganhos .gasto { background-color: #FFEBEE !important; }
    .tabela-ganhos .positivo { color: #2E7D32; font-weight: 700; }
    .tabela-ganhos .negativo { color: #C62828; font-weight: 700; }
    </style>
    <table class="tabela-ganhos">
    <thead>
    <tr>
        <th>Data</th>
        <th>Dia da Semana</th>
        <th>Km Executado</th>
        <th>Diária</th>
        <th>Qtd. Pacotes</th>
        <th>Qtd. Entregas</th>
        <th>% Entrega</th>
        <th>Bônus</th>
        <th>Total Ganhos</th>
        <th>Gasolina</th>
        <th>Almoço</th>
        <th>Besteira</th>
        <th>Carro</th>
        <th>Total Gastos</th>
        <th>Lucro Diário</th>
    </tr>
    </thead>
    <tbody>
    """

    for _, row in fmt.iterrows():
        lucro_cls = "positivo" if row["lucro"] >= 0 else "negativo"
        html += f"""<tr>
            <td class="text-left">{row['data'].strftime('%d/%m/%Y') if pd.notna(row['data']) else ''}</td>
            <td class="text-left">{row['dia_semana']}</td>
            <td class="numero">{int(row['km']) if pd.notna(row['km']) else ''}</td>
            <td class="numero">R$ {fmt_val(row['diaria'])}</td>
            <td class="numero">{int(row['pacotes']) if pd.notna(row['pacotes']) else ''}</td>
            <td class="numero">{int(row['entregas']) if pd.notna(row['entregas']) else ''}</td>
            <td class="numero">{fmt_val(row['pct_entrega']*100 if pd.notna(row['pct_entrega']) else 0, dec=0)}%</td>
            <td class="numero">R$ {fmt_val(row['bonus'])}</td>
            <td class="numero ganho">R$ {fmt_val(row['total_ganho'])}</td>
            <td class="numero">R$ {fmt_val(row['gasolina'], dec=2)}</td>
            <td class="numero">R$ {fmt_val(row['almoco'], dec=2)}</td>
            <td class="numero">R$ {fmt_val(row['besteira'], dec=2)}</td>
            <td class="numero">R$ {fmt_val(row['carro'], dec=2)}</td>
            <td class="numero gasto">R$ {fmt_val(row['total_gasto'], dec=2)}</td>
            <td class="numero {lucro_cls}">R$ {fmt_val(row['lucro'], dec=2)}</td>
        </tr>"""

    html += "</tbody></table>"

    st_html(html, height=600, scrolling=True)

st.divider()
if st.button("Exportar para Excel"):
    with pd.ExcelWriter("Planilha_de_Ganhos_Exportada.xlsx", engine="openpyxl") as writer:
        for mes, d in st.session_state["dados"].items():
            calcular_df(d).to_excel(writer, sheet_name=mes, index=False)
    st.success("Planilha exportada com sucesso!")
